import openpyxl
import os
import sys

os.chdir('/Users/Jacob/Desktop/SSIC')

from openpyxl import Workbook
from openpyxl import load_workbook

wb = load_workbook('ComplianceForge - Cybersecurity Maturity Model Certification (CMMC) v1.02 Requirements Matrix.xlsx')
ws = wb['CMMC v1.02']




CMMC = []
CISCSC = []

def noDupe(Lis):
    res = []
    for var in Lis:
        if var not in res:
            res.append(var)
    return (res)

for n in range(3, 258):
    
    if str(ws.cell(row=n, column=17).value) != 'None':
        temp = []
        temp.append(str(ws.cell(row=n, column=3).value))
        splitLst = str(ws.cell(row=n, column=17).value).split()
        splitLst = noDupe(splitLst)

        for var in splitLst:
            if var not in CISCSC:
                CISCSC.append(var)
        


        for i in splitLst:
            temp.append(i)
        CMMC.append(temp)
    

y = sorted(CISCSC, key=lambda x: float(x))

def toString(s):
    str1 = ', '
    return (str1.join(s))

for val in y:
    temp = []
    temp.append(val)
    for cmmcval in CMMC:
        if val in cmmcval:
            temp.append(cmmcval[0])
    print(toString(temp))
            
            



        
        
    


        


