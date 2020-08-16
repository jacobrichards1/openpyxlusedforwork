import openpyxl
import os
import sys

os.chdir('/Users/Jacob/Desktop/SSIC')

from openpyxl import Workbook
from openpyxl import load_workbook

wb = load_workbook('CIS-NIST171-NIST53-ISO.xlsx')
ws = wb['Sheet1']

maxGroup = 'Group 1'
NIST = []
NIST171 =[]
NIST53 = []
ISO = [] 

for line in sys.stdin:
    n = int(line)
    if n == -1:
        break
    for i in range (2, 173):
        if n == ws.cell(row=i, column=1).value:
            ImpGroup = ws.cell(row=i, column=3).value
            if ImpGroup > maxGroup:
                maxGroup = ImpGroup

            if str(ws.cell(row=i, column=4).value) != 'None':
                tempNIST = str(ws.cell(row=i, column=4).value)
                if str(ws.cell(row=i, column=4).value).find(',') >= 0:
                    tempNIST = tempNIST.replace(',','')
                    temptemp = tempNIST.split()
                    for va in temptemp:
                        NIST.append(va)
                else:
                    NIST.append(tempNIST)

            if str(ws.cell(row=i, column=5).value) != 'None':
                tempNIST171 = str(ws.cell(row=i, column=5).value)
                if str(ws.cell(row=i, column=5).value).find(',') >= 0:
                    tempNIST171 = tempNIST171.replace(',','')
                    temptemp = tempNIST171.split()
                    for va in temptemp:
                        NIST171.append(va)
                else:
                    NIST171.append(tempNIST171)

            if str(ws.cell(row=i, column=6).value) != 'None':
                tempNIST53 = str(ws.cell(row=i, column=6).value)
                if str(ws.cell(row=i, column=6).value).find(',') >= 0:
                    tempNIST53 = tempNIST53.replace(',','')
                    temptemp = tempNIST53.split()
                    for va in temptemp:
                        NIST53.append(va)
                else:
                    NIST53.append(tempNIST53)

            if str(ws.cell(row=i, column=7).value) != 'None':
                tempISO = str(ws.cell(row=i, column=7).value)
                if str(ws.cell(row=i, column=7).value).find(',') >= 0:
                    tempISO = tempISO.replace(',','')
                    temptemp = tempISO.split()
                    for va in temptemp:
                        ISO.append(va)
                else:
                    ISO.append(tempISO)


def noDupe(Lis):
    res = []
    for var in Lis:
        if var not in res:
            res.append(var)
    return (res)


def toString(s):
    str1 = ', '
    return (str1.join(s))


print(maxGroup)
NIST = noDupe(NIST)
NIST.sort()        
print(toString(NIST))
NIST171 = noDupe(NIST171)
NIST171.sort()
print(toString(NIST171))
NIST53 = noDupe(NIST53)
NIST53.sort()
print(toString(NIST53))
ISO = noDupe(ISO)
ISO.sort()
print(toString(ISO))


            


