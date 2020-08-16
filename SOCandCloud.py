import openpyxl
import os
import sys

os.chdir('/Users/Jacob/Desktop/SSIC')

from openpyxl import Workbook
from openpyxl import load_workbook

wb = load_workbook('AuditScripts-CIS-Controls-Master-Mappings-v7.1c.xlsx')
wsSOC = wb['AICPA SOC2 and SOC3 TSC 2017']
wsCloud = wb['CSA CCM v3']

SOC = []
Cloud = []

def noDupe(Lis):
    res = []
    for var in Lis:
        if var not in res:
            res.append(var)
    return (res)

def toString(s):
    str1 = ', '
    return (str1.join(s))

for r in  range(5, 194):
    for c in range(4, 136):
        if str(wsCloud.cell(row=r, column=c).value) == 'X':
            temp = []
            temp.append(str(wsCloud.cell(row=r, column=2).value))
            temp.append(str(wsCloud.cell(row=4, column=c).value))
            Cloud.append(temp)

for r in  range(5, 194):
    for c in range(4, 50):
        if str(wsSOC.cell(row=r, column=c).value) == 'X':
            temp = []
            temp.append(str(wsSOC.cell(row=r, column=2).value))
            temp.append(str(wsSOC.cell(row=4, column=c).value))
            SOC.append(temp)

CloudRes = []
SOCRes = []

for val in Cloud:
    temp = []
    if val[0] not in CloudRes:
        temp.append(val[0])
    for counter in Cloud:
        if val[0] == counter[0]:
            temp.append(counter[1])
            CloudRes.append(temp)



for var in SOC:
    temp = []
    if var[0] not in SOCRes:
        temp.append(var[0])
    for counter in SOC:
        if var[0] == counter[0]:
            temp.append(counter[1])
            SOCRes.append(temp)


    
CloudRes = noDupe(CloudRes)
for i in CloudRes:
    print (toString(i))

print(' ')
print(' ')
print(' ')

SOCRes = noDupe(SOCRes)
for i in SOCRes:
    print (toString(i))

    



