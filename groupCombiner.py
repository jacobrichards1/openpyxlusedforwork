import openpyxl
import os
import sys

os.chdir('/Users/Jacob/Desktop/SSIC')

from openpyxl import Workbook
from openpyxl import load_workbook

wb = load_workbook('out.xlsx')
ws = wb.active

a1 = ws.cell(row=4, column=7).value
a2 = ws.cell(row=13, column=7).value
a3 = ws.cell(row=21, column=7).value
a4 = ws.cell(row=29, column=7).value
a5 = ws.cell(row=37, column=7).value
a6 = ws.cell(row=43, column=7).value
a7 = ws.cell(row=52, column=7).value
a8 = ws.cell(row=61, column=7).value
a9 = ws.cell(row=68, column=7).value
a10 = ws.cell(row=73, column=7).value
a11 = ws.cell(row=79, column=7).value
a12 = ws.cell(row=87, column=7).value
a13 = ws.cell(row=99, column=7).value
a14 = ws.cell(row=108, column=7).value
a15 = ws.cell(row=117, column=7).value
a16 = ws.cell(row=129, column=7).value
a17 = ws.cell(row=139, column=7).value
a18 = ws.cell(row=150, column=7).value
a19 = ws.cell(row=159, column=7).value
a20 = ws.cell(row=167, column=7).value

fullList = []

def valuesFrom(i):
    switcher={
        1:a1,
        2:a2,
        3:a3,
        4:a4,
        5:a5,
        6:a6,
        7:a7,
        8:a8,
        9:a9,
        10:a10,
        11:a11,
        12:a12,
        13:a13,
        14:a14,
        15:a15,
        16:a16,
        17:a17,
        18:a18,
        19:a19,
        20:a20
    }
    return switcher.get(i, "")

for line in sys.stdin:
    n = int(line)
    if n == -1:
        break
    for var in valuesFrom(n).split():
        fullList.append(var)

def toString(s):
    str1 = ' '
    return (str1.join(s))

res = []
for i in fullList:
    if i not in res:
        res.append(i)
res.sort()
toString(res)
print(toString(res))




    




    




