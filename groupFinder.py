import openpyxl
import os

os.chdir('/Users/Jacob/Desktop/SSIC')

from openpyxl import Workbook
from openpyxl import load_workbook

wb = load_workbook('CIS-Controls-Version-7.1-Implementation-Groups.xlsx')
ws = wb['All CIS Controls & Sub-Controls']

wb2 = load_workbook('out.xlsx')
ws2 = wb2.active

count = 0
curRow = 1
curCol = 1
rowNum = 0
group = ''

for rowI in range(4, 213):
    for i in range(2, 11):
        cellObj = ws.cell(row=rowI, column=i)
        if cellObj.value == 'X':
            count += 1
    
    if count == 1:
        group = 'Group 3'
        rowNum = ws.cell(row=rowI, column=3).value
    elif count == 3:
        group = 'Group 1'
        rowNum = ws.cell(row=rowI, column=3).value
    elif count == 2:
        group = 'Group 2'
        rowNum = ws.cell(row=rowI, column=3).value

    if count != 0:
        ws2.cell(row=curRow, column=curCol).value = ws.cell(row=rowI, column=3).value
        ws2.cell(row=curRow, column=curCol+1).value = group
        curRow += 1

    
    count = 0
    group = ''

wb2.save(filename='out.xlsx')


